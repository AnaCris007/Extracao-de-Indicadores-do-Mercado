"""Planilha de premissas macro para valuation: IPCA, Selic, PIB, Ouro e Prata.

Fontes oficiais/gratuitas, sem chave de API:
    - IPCA, histórico              -> IBGE / SIDRA (tabela 1737, variação
                                      acumulada no ano), 1 linha por ano fechado
    - Selic, histórico             -> Banco Central (SGS série 432, meta Selic
                                      realizada), 1 linha por ano fechado
    - PIB, histórico                -> IBGE / SIDRA (tabela 6784, Contas
                                      Nacionais - variação real anual), 1
                                      linha por ano já publicado pelo IBGE
                                      (esse dado sai com defasagem de ~1-2
                                      anos - é normal faltar o(s) ano(s) mais
                                      recente(s), viram projeção até lá)
    - IPCA/Selic/PIB, projeção     -> Focus / Banco Central (Olinda OData),
                                      mediana da pesquisa mais recente, 1
                                      linha por ano ainda sem dado oficial
                                      fechado (curva de projeção, não o
                                      histórico de como a previsão mudou)
    - Ouro e Prata, histórico      -> Yahoo Finance (tickers GC=F e SI=F),
                                      cotação em USD/onça troy, como
                                      negociado no mercado internacional
    - USD/BRL                      -> Banco Central (SGS série 1, câmbio
                                      PTAX venda), só a cotação mais atual
                                      (1 linha, no ano corrente - não é
                                      série histórica como Ouro/Prata)

A planilha final fica em formato largo (`pivotar_tabela()`): 1 linha por
indicador, 1 coluna por ano - ano mais antigo à esquerda, mais recente à
direita. O ano corrente (`ANO_TRIMESTRAL`) ganha 4 colunas extras
(`adicionar_trimestres()`) - 1T/2T/3T/4T - pros indicadores com dado
sub-anual disponível (IPCA, Ouro, Prata, USD/BRL).

As funções de coleta (`get_*`), `montar_tabela()` e `pivotar_tabela()`
também são reaproveitadas pelo app visual em `scripts/app.py`
(`streamlit run scripts/app.py`), que deixa escolher indicadores/período
numa tela em vez de editar o código.

pip install pandas requests yfinance openpyxl

Uso:
    python scripts/premissas_macro.py
"""

from concurrent.futures import ThreadPoolExecutor
from datetime import date, timedelta
from urllib.parse import quote

import pandas as pd
import requests
import yfinance as yf
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HOJE = date.today()
ANO_INICIO_PADRAO = HOJE.year - 7  # janela padrão de histórico
ANO_FIM_PADRAO = HOJE.year + 4  # até onde a projeção Focus vai por padrão
ANO_TRIMESTRAL = HOJE.year  # único ano que ganha colunas 1T/2T/3T/4T na planilha final
OUTPUT = "data/premissas_valuation.xlsx"

# Indicadores "macro" (1 linha/ano, Histórico até fechar + Projeção Focus
# depois disso) e o nome que a API do BCB Focus usa para cada um.
INDICADORES_MACRO = {
    "IPCA": "% a.a. (variação acumulada no ano)",
    "Selic": "% a.a. (taxa básica de juros)",
    "PIB": "% a.a. (variação do PIB no ano)",
}
INDICADOR_FOCUS = {"IPCA": "IPCA", "Selic": "Selic", "PIB": "PIB Total"}

# Commodities (histórico diário, sem projeção) e seu ticker na Yahoo Finance.
# Cotadas nativamente em USD/onça troy - não convertidas para BRL.
INDICADORES_COMMODITY = {"Ouro": "GC=F", "Prata": "SI=F"}

# Câmbio (histórico diário, sem projeção).
INDICADOR_CAMBIO = "USD/BRL"

TODOS_INDICADORES = list(INDICADORES_MACRO) + list(INDICADORES_COMMODITY) + [INDICADOR_CAMBIO]

# De onde vem o valor de cada indicador - vira a coluna "Fonte" na planilha
# final. Nos indicadores macro, histórico e projeção vêm de fontes
# diferentes (IBGE/BCB vs. Focus), por isso os dois aparecem juntos.
FONTES = {
    "IPCA": "IBGE/SIDRA (histórico) + Focus/BCB (projeção)",
    "Selic": "BCB SGS 432 (histórico) + Focus/BCB (projeção)",
    "PIB": "IBGE/SIDRA (histórico) + Focus/BCB (projeção)",
    "Ouro": "Yahoo Finance (GC=F)",
    "Prata": "Yahoo Finance (SI=F)",
    INDICADOR_CAMBIO: "BCB SGS 1 (PTAX)",
}


def get_ipca_historico() -> pd.DataFrame:
    """IPCA realizado, variação acumulada no ano (%), 1 linha por ano - IBGE/SIDRA.

    Variável 69 da tabela 1737 é o acumulado desde janeiro; pegando só o
    valor de dezembro de cada ano, obtemos o IPCA fechado do ano - no
    mesmo formato (% a.a.) da projeção do Focus, pra ficar comparável.
    """
    url = "https://apisidra.ibge.gov.br/values/t/1737/n1/all/v/69/p/all"
    rows = requests.get(url, timeout=30).json()[1:]  # 1ª linha é cabeçalho
    df = pd.DataFrame(rows)
    df = df[df["D3C"].str.endswith("12") & (df["V"] != "...")]  # só dezembro = ano fechado
    return pd.DataFrame(
        {
            "Data": pd.to_datetime(df["D3C"], format="%Y%m"),
            "Indicador": "IPCA",
            "Tipo": "Histórico",
            "Valor": df["V"].astype(float),
            "Unidade": INDICADORES_MACRO["IPCA"],
            "AnoReferencia": df["D3C"].str[:4].astype(int),
        }
    ).reset_index(drop=True)


def get_ipca_trimestral(ano: int) -> pd.DataFrame:
    """IPCA acumulado por trimestre civil do `ano` pedido - IBGE/SIDRA.

    Variável 2263 é "acumulado em 3 meses" (janela móvel, fecha todo mês);
    pegando só os meses 03/06/09/12, essa janela coincide exatamente com
    o trimestre civil (jan-fev-mar, abr-mai-jun, ...). Só quarters já
    fechados vêm na resposta - a API do IBGE simplesmente omite os meses
    ainda não publicados, sem "..." nem erro.
    """
    meses = [f"{ano}{m:02d}" for m in (3, 6, 9, 12)]
    url = f"https://apisidra.ibge.gov.br/values/t/1737/n1/all/v/2263/p/{','.join(meses)}"
    rows = requests.get(url, timeout=30).json()[1:]  # 1ª linha é cabeçalho
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Data", "Indicador", "Tipo", "Valor", "Unidade", "Trimestre"])
    df = df[df["V"] != "..."]
    mes_para_trimestre = {"03": 1, "06": 2, "09": 3, "12": 4}
    return pd.DataFrame(
        {
            "Data": pd.to_datetime(df["D3C"], format="%Y%m"),
            "Indicador": "IPCA",
            "Tipo": "Histórico",
            "Valor": df["V"].astype(float),
            "Unidade": "% (acumulado no trimestre)",
            "Trimestre": df["D3C"].str[-2:].map(mes_para_trimestre),
        }
    ).reset_index(drop=True)


def get_pib_historico() -> pd.DataFrame:
    """PIB realizado, variação real anual (%), 1 linha por ano - IBGE/SIDRA
    (Contas Nacionais Anuais, tabela 6784).

    O IBGE fecha as Contas Nacionais Anuais com defasagem de ~1-2 anos, ao
    contrário do IPCA/Selic - então o(s) ano(s) mais recente(s) simplesmente
    não aparecem aqui ainda; `montar_tabela()` completa a lacuna com a
    projeção Focus.
    """
    url = "https://apisidra.ibge.gov.br/values/t/6784/n1/all/v/9810/p/all"
    rows = requests.get(url, timeout=30).json()[1:]  # 1ª linha é cabeçalho
    df = pd.DataFrame(rows)
    df = df[df["V"] != "..."]
    return pd.DataFrame(
        {
            "Data": pd.to_datetime(df["D3C"] + "-12-31"),
            "Indicador": "PIB",
            "Tipo": "Histórico",
            "Valor": df["V"].astype(float),
            "Unidade": INDICADORES_MACRO["PIB"],
            "AnoReferencia": df["D3C"].astype(int),
        }
    ).reset_index(drop=True)


def get_sgs_serie(codigo: int, ano_inicio: int) -> pd.DataFrame:
    """Busca uma série diária do SGS/BCB (colunas Data/Valor), quebrando a
    consulta em janelas de até 10 anos.

    A própria API do Banco Central recusa (HTTP 406) qualquer consulta de
    série diária com janela maior que 10 anos - então pedir, por exemplo,
    2015 a 2030 numa chamada só falha. Aqui isso é escondido do resto do
    código: quem chama só pensa em "ano inicial", sem se preocupar com o
    limite.
    """
    url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}/dados"
    inicio = date(ano_inicio, 1, 1)
    partes = []
    while inicio <= HOJE:
        fim_janela = min(inicio + timedelta(days=3650), HOJE)  # ~9 anos e 11 meses, sempre < 10 anos
        params = {"formato": "json", "dataInicial": inicio.strftime("%d/%m/%Y"), "dataFinal": fim_janela.strftime("%d/%m/%Y")}

        payload = None
        for tentativa in range(2):  # a API do BCB às vezes falha/demora de forma transitória
            try:
                payload = requests.get(url, params=params, timeout=15).json()
                break
            except (requests.RequestException, ValueError):
                if tentativa == 1:
                    print(f"aviso: falha ao buscar série {codigo} de {inicio} a {fim_janela}, pulando janela")

        if isinstance(payload, list) and payload:
            partes.append(pd.DataFrame(payload))
        inicio = fim_janela + timedelta(days=1)

    if not partes:
        return pd.DataFrame({"Data": pd.Series(dtype="datetime64[ns]"), "Valor": pd.Series(dtype="float64")})
    df = pd.concat(partes, ignore_index=True)
    df["Data"] = pd.to_datetime(df["data"], format="%d/%m/%Y")
    df["Valor"] = df["valor"].astype(float)
    return df[["Data", "Valor"]].sort_values("Data").reset_index(drop=True)


def get_selic_historico(ano_inicio: int) -> pd.DataFrame:
    """Selic (meta Copom) realizada, 1 linha por ano já fechado - BCB, SGS série 432.

    A série 432 repete o valor vigente em todo dia útil (a meta só muda em
    reunião do Copom); aqui pegamos o último valor de cada ano - a Selic
    com que aquele ano realmente fechou. Só anos anteriores ao corrente
    entram aqui: o ano corrente em diante já vem do Focus (get_focus), que
    é projeção por definição, ainda não fechou.
    """
    df = get_sgs_serie(432, ano_inicio)
    df = df[df["Data"].dt.year < HOJE.year]

    ultimo_por_ano = df.sort_values("Data").groupby(df["Data"].dt.year).last()
    return pd.DataFrame(
        {
            "Data": ultimo_por_ano["Data"],
            "Indicador": "Selic",
            "Tipo": "Histórico",
            "Valor": ultimo_por_ano["Valor"],
            "Unidade": INDICADORES_MACRO["Selic"],
            "AnoReferencia": ultimo_por_ano.index,
        }
    ).reset_index(drop=True)


def anos_sem_historico(historico: pd.DataFrame, ano_inicio: int, ano_fim: int) -> list[int]:
    """Anos entre o último dado oficial fechado e o fim da janela pedida.

    O IPCA/Selic fecham o ano anterior; o PIB do IBGE atrasa mais (só sai
    ~1-2 anos depois) - por isso cada indicador pode ter uma lacuna de
    tamanho diferente, calculada aqui em vez de fixa. Se não há histórico
    nenhum no período (ex.: usuário pediu só anos futuros), cobre tudo com
    projeção.
    """
    if historico.empty:
        return list(range(ano_inicio, ano_fim + 1))
    ultimo_ano_fechado = int(historico["AnoReferencia"].max())
    return list(range(max(ultimo_ano_fechado + 1, ano_inicio), ano_fim + 1))


def get_focus(indicador: str, nome: str, unidade: str, anos: list[int]) -> pd.DataFrame:
    """Curva de projeção Focus/BCB: só a pesquisa mais recente por ano-alvo.

    Uma linha por ano em `anos`, cada uma com a mediana da última pesquisa
    disponível pra aquele ano - não o histórico de como a previsão mudou ao
    longo do tempo (isso existe, mas não serve pra alimentar um modelo de
    valuation: o que importa é o consenso de mercado *atual* para cada ano
    futuro).

    Uma chamada HTTP por ano - feitas em paralelo (`ThreadPoolExecutor`)
    pra não multiplicar o tempo de espera pelo número de anos. Timeout e
    tentativas propositalmente curtos: se a API do BCB estiver
    inacessível (rede/firewall bloqueando `bcb.gov.br`, não é raro em
    rede corporativa/institucional), é melhor falhar rápido e seguir sem
    a projeção daquele ano do que travar o app por minutos.
    """

    def buscar_ano(ano: int) -> dict | None:
        filtro = f"Indicador eq '{indicador}' and DataReferencia eq '{ano}' and baseCalculo eq 0"
        params = {"$filter": filtro, "$orderby": "Data desc", "$top": "1", "$format": "json"}
        # A API do BCB só aceita espaço como %20 (não '+'), por isso o encode manual.
        query = "&".join(f"{k}={quote(v)}" for k, v in params.items())
        url = f"https://olinda.bcb.gov.br/olinda/servico/Expectativas/versao/v1/odata/ExpectativasMercadoAnuais?{query}"

        payload = None
        for tentativa in range(2):
            try:
                payload = requests.get(url, timeout=15).json()["value"]
                break
            except (requests.RequestException, ValueError):
                if tentativa == 1:
                    print(f"aviso: falha ao buscar projeção Focus de {indicador} para {ano}, pulando ano")

        if not payload:  # nenhuma pesquisa registrada pra esse ano ainda, ou falha persistente
            return None
        registro = payload[0]
        return {
            "Data": pd.to_datetime(registro["Data"]),
            "Indicador": nome,
            "Tipo": "Projeção",
            "Valor": float(registro["Mediana"]),
            "Unidade": unidade,
            "AnoReferencia": ano,
        }

    if not anos:
        return pd.DataFrame(columns=["Data", "Indicador", "Tipo", "Valor", "Unidade", "AnoReferencia"])

    with ThreadPoolExecutor(max_workers=min(8, len(anos))) as executor:
        resultados = executor.map(buscar_ano, anos)
        linhas = [linha for linha in resultados if linha is not None]
    return pd.DataFrame(linhas)


def get_usd_brl() -> pd.DataFrame:
    """Câmbio USD/BRL (PTAX venda) - só a cotação mais atual, Banco Central SGS série 1.

    Diferente de Ouro/Prata (série histórica completa), aqui só interessa
    o índice do dia - 1 única linha, no ano corrente. `/dados/ultimos/1` é
    uma chamada rápida e leve, em vez de baixar anos de série diária só
    pra usar o último valor.
    """
    url = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.1/dados/ultimos/1"
    payload = None
    for tentativa in range(2):
        try:
            payload = requests.get(url, params={"formato": "json"}, timeout=15).json()
            break
        except (requests.RequestException, ValueError):
            if tentativa == 1:
                print("aviso: falha ao buscar cotação atual do USD/BRL")

    if not payload:
        return pd.DataFrame(columns=["Data", "Indicador", "Tipo", "Valor", "Unidade", "AnoReferencia"])

    registro = payload[0]
    data = pd.to_datetime(registro["data"], format="%d/%m/%Y")
    return pd.DataFrame(
        [
            {
                "Data": data,
                "Indicador": INDICADOR_CAMBIO,
                "Tipo": "Histórico",
                "Valor": float(registro["valor"]),
                "Unidade": "BRL",
                "AnoReferencia": data.year,
            }
        ]
    )


def get_commodity(ticker: str, nome: str, ano_inicio: int) -> pd.DataFrame:
    """Cotação histórica de fechamento - Yahoo Finance, em USD/onça troy.

    Diferente de get_ipca/get_focus, aqui não há uma URL explícita: a
    biblioteca `yfinance` monta e chama a API do Yahoo Finance
    internamente (endpoint privado, não documentado publicamente -
    https://query1.finance.yahoo.com/v8/finance/chart/<ticker>), então o
    HTTP fica escondido dentro de `yf.Ticker(ticker).history(...)`.

    GC=F e SI=F (futuros de ouro/prata da COMEX) são cotados nativamente
    em dólar por onça troy (1 onça troy = 31,1035 g) - mantido em USD, como
    negociado no mercado internacional, sem converter para BRL.

    O Yahoo Finance às vezes falha de forma transitória e devolve uma
    tabela vazia (sem levantar exceção) - `hist.index` nesse caso não é
    um `DatetimeIndex`, então tenta de novo antes de desistir e pular o
    indicador, em vez de deixar o resto da coleta quebrar por causa disso.
    """
    hist = yf.Ticker(ticker).history(start=date(ano_inicio, 1, 1), interval="1d")
    if hist.empty:
        hist = yf.Ticker(ticker).history(start=date(ano_inicio, 1, 1), interval="1d")

    if hist.empty:
        print(f"aviso: Yahoo Finance não retornou cotação para {nome} ({ticker}), pulando indicador")
        return pd.DataFrame(columns=["Data", "Indicador", "Tipo", "Valor", "Unidade", "AnoReferencia"])

    dates = hist.index.tz_localize(None).astype("datetime64[ns]")
    return pd.DataFrame(
        {
            "Data": dates,
            "Indicador": nome,
            "Tipo": "Histórico",
            "Valor": hist["Close"].round(2).to_numpy(),
            "Unidade": "USD/onça troy",
            "AnoReferencia": dates.year,
        }
    ).sort_values("Data").reset_index(drop=True)


def pivotar_tabela(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, str]]:
    """Reformata a tabela longa (1 linha por Data/Indicador) em formato
    largo: 1 linha por indicador, 1 coluna por ano - ano mais antigo à
    esquerda, mais recente à direita.

    Indicadores anuais (IPCA/Selic/PIB) já têm só um valor por ano.
    Indicadores diários (Ouro, Prata, USD/BRL) usam o último valor
    observado em cada ano - mesmo critério de "ano fechado" já usado no
    histórico da Selic (`get_selic_historico`).

    Devolve `(valores, tipos, unidades)`: `valores` e `tipos` no mesmo
    formato largo (índice = Indicador, colunas = Ano), usados pra escrever
    a planilha e colorir Histórico/Projeção; `unidades` mapeia indicador
    -> unidade, usado pra formatar cada linha (%, USD, R$).
    """
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), {}

    agrupado = df.sort_values("Data").groupby(["Indicador", "AnoReferencia"])
    valores = agrupado["Valor"].last().unstack("AnoReferencia")
    tipos = agrupado["Tipo"].last().unstack("AnoReferencia")

    ordem = [nome for nome in TODOS_INDICADORES if nome in valores.index]
    anos = sorted(valores.columns)
    valores = valores.reindex(index=ordem, columns=anos)
    tipos = tipos.reindex(index=ordem, columns=anos)
    unidades = df.drop_duplicates("Indicador").set_index("Indicador")["Unidade"].to_dict()
    return valores, tipos, unidades


def adicionar_trimestres(valores: pd.DataFrame, tipos: pd.DataFrame, df: pd.DataFrame, ano: int) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Quebra a coluna do `ano` pedido em 4 colunas trimestrais
    (1T-{ano} .. 4T-{ano}), inseridas antes da coluna anual - só pros
    indicadores com granularidade sub-anual disponível:

    - Ouro/Prata: último valor observado em cada trimestre (mesmo
      critério do fechamento anual), a partir da série diária já
      coletada em `df` (`montar_tabela`) - sem nenhuma chamada extra.
    - USD/BRL: só há a cotação atual coletada (`get_usd_brl`); ela cai
      no trimestre da sua própria data, os outros 3 ficam em branco -
      não busca histórico só pra preencher a planilha.
    - IPCA: 1 chamada extra ao IBGE (`get_ipca_trimestral`), só quando
      `ano` está entre as colunas de `valores`.
    - Selic/PIB: sem fonte de dado trimestral hoje - colunas em branco.

    Se `ano` não estiver entre as colunas de `valores` (ex.: usuário
    pediu um período que não inclui esse ano), devolve tudo sem
    alteração.
    """
    if valores.empty or ano not in valores.columns:
        return valores, tipos

    colunas_tri = [f"{t}T-{ano}" for t in (1, 2, 3, 4)]
    valores_tri = pd.DataFrame(index=valores.index, columns=colunas_tri, dtype=float)
    tipos_tri = pd.DataFrame(index=valores.index, columns=colunas_tri, dtype=object)

    diario = df[(df["AnoReferencia"] == ano) & df["Indicador"].isin(["Ouro", "Prata", INDICADOR_CAMBIO])]
    if not diario.empty:
        agrupado = diario.assign(Trimestre=diario["Data"].dt.quarter).sort_values("Data").groupby(["Indicador", "Trimestre"])
        for (indicador, trimestre), valor in agrupado["Valor"].last().items():
            if indicador in valores_tri.index:
                col = f"{trimestre}T-{ano}"
                valores_tri.loc[indicador, col] = valor
                tipos_tri.loc[indicador, col] = agrupado["Tipo"].last()[(indicador, trimestre)]

    if "IPCA" in valores.index:
        for _, linha in get_ipca_trimestral(ano).iterrows():
            col = f"{int(linha['Trimestre'])}T-{ano}"
            valores_tri.loc["IPCA", col] = linha["Valor"]
            tipos_tri.loc["IPCA", col] = linha["Tipo"]

    posicao = list(valores.columns).index(ano)
    ordem_colunas = list(valores.columns[:posicao]) + colunas_tri + list(valores.columns[posicao:])
    valores = pd.concat([valores, valores_tri], axis=1)[ordem_colunas]
    tipos = pd.concat([tipos, tipos_tri], axis=1)[ordem_colunas]
    return valores, tipos


def montar_tabela(indicadores: list[str], ano_inicio: int, ano_fim: int) -> pd.DataFrame:
    """Coleta só os indicadores/anos pedidos e devolve a tabela consolidada.

    `indicadores` é qualquer subconjunto de `TODOS_INDICADORES`. Usada tanto
    pelo `main()` (CLI, roda com tudo) quanto pelo `app.py` (Streamlit, roda
    só com o que o usuário selecionou na tela).
    """
    tabelas: list[pd.DataFrame] = []

    for nome in indicadores:
        if nome in INDICADORES_MACRO:
            historico = {
                "IPCA": get_ipca_historico,
                "Selic": lambda: get_selic_historico(ano_inicio),
                "PIB": get_pib_historico,
            }[nome]()
            anos = anos_sem_historico(historico, ano_inicio, ano_fim)
            projecao = get_focus(INDICADOR_FOCUS[nome], nome, INDICADORES_MACRO[nome], anos)
            tabelas += [historico, projecao]
        elif nome in INDICADORES_COMMODITY:
            tabelas.append(get_commodity(INDICADORES_COMMODITY[nome], nome, ano_inicio))
        elif nome == INDICADOR_CAMBIO:
            tabelas.append(get_usd_brl())

    tabelas = [t for t in tabelas if not t.empty]
    if not tabelas:
        return pd.DataFrame(columns=["Data", "Indicador", "Tipo", "Valor", "Unidade", "AnoReferencia"])

    df = pd.concat(tabelas, ignore_index=True)
    df = df[(df["AnoReferencia"] >= ano_inicio) & (df["AnoReferencia"] <= ano_fim)]
    return df.sort_values(["Indicador", "Data"]).reset_index(drop=True)


def formato_numero(unidade: str) -> str:
    """Formato numérico do Excel apropriado pra unidade do indicador."""
    if "%" in unidade:
        return '0.00"%"'
    if unidade == "BRL":  # câmbio USD/BRL
        return '#,##0.0000" R$"'
    return '#,##0.00" USD"'  # commodities (USD/onça troy)


def montar_planilha(valores: pd.DataFrame, unidades: dict[str, str]) -> pd.DataFrame:
    """DataFrame pronto pra `to_excel`: colunas `Unidade` e `Fonte`
    inseridas logo depois do índice (Indicador), antes das colunas de ano."""
    saida = valores.copy()
    saida.insert(0, "Fonte", [FONTES.get(indicador, "") for indicador in saida.index])
    saida.insert(0, "Unidade", [unidades.get(indicador, "") for indicador in saida.index])
    return saida


# Layout fixo da planilha final: A = Indicador (índice), B = Unidade,
# C = Fonte, D em diante = anos.
COL_UNIDADE = 2
COL_FONTE = 3
COL_PRIMEIRO_ANO = 4


def formatar_planilha(caminho: str, valores: pd.DataFrame, tipos: pd.DataFrame, unidades: dict[str, str]) -> None:
    """Aplica formatação visual ao .xlsx já salvo: cabeçalho, uma linha por
    indicador com cor por Tipo (Histórico/Projeção) e formato numérico por
    linha (%, USD, R$), largura de coluna.

    Reaberto depois do `to_excel` de propósito - separa "gerar o dado" de
    "deixar bonito", em vez de misturar estilo com a lógica de coleta.
    """
    from openpyxl import load_workbook

    wb = load_workbook(caminho)
    ws = wb["Premissas"]

    cor_header = PatternFill("solid", fgColor="1F4E78")
    cor_header_trimestre = PatternFill("solid", fgColor="000000")  # destaca colunas 1T/2T/3T/4T
    cor_historico = PatternFill("solid", fgColor="E2EFDA")  # verde claro
    cor_projecao = PatternFill("solid", fgColor="FCE4D6")  # laranja claro
    centralizado = Alignment(horizontal="center", vertical="center")

    n_indicadores, n_anos = valores.shape
    ultima_coluna = COL_PRIMEIRO_ANO + n_anos - 1

    for c in range(1, ultima_coluna + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        eh_trimestre = c >= COL_PRIMEIRO_ANO and isinstance(valores.columns[c - COL_PRIMEIRO_ANO], str)
        cell.fill = cor_header_trimestre if eh_trimestre else cor_header
        cell.alignment = centralizado

    for r, indicador in enumerate(valores.index, start=2):
        label = ws.cell(r, 1)
        label.font = Font(bold=True)
        label.alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(r, COL_UNIDADE).alignment = centralizado
        ws.cell(r, COL_FONTE).alignment = Alignment(horizontal="left", vertical="center")

        formato = formato_numero(unidades.get(indicador, ""))
        for c, ano in enumerate(valores.columns, start=COL_PRIMEIRO_ANO):
            valor = valores.loc[indicador, ano]
            if pd.isna(valor):
                continue
            cell = ws.cell(r, c)
            cell.alignment = centralizado
            cell.number_format = formato
            cell.fill = cor_historico if tipos.loc[indicador, ano] == "Histórico" else cor_projecao

    ws.column_dimensions["A"].width = max((len(str(i)) for i in valores.index), default=8) + 4
    ws.column_dimensions[get_column_letter(COL_UNIDADE)].width = max((len(str(u)) for u in unidades.values()), default=8) + 4
    ws.column_dimensions[get_column_letter(COL_FONTE)].width = max((len(str(f)) for f in FONTES.values()), default=8) + 4
    for c in range(COL_PRIMEIRO_ANO, ultima_coluna + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = ws.cell(2, COL_PRIMEIRO_ANO).coordinate
    wb.save(caminho)


def main() -> None:
    df = montar_tabela(TODOS_INDICADORES, ANO_INICIO_PADRAO, ANO_FIM_PADRAO)
    valores, tipos, unidades = pivotar_tabela(df)
    valores, tipos = adicionar_trimestres(valores, tipos, df, ANO_TRIMESTRAL)
    montar_planilha(valores, unidades).to_excel(OUTPUT, sheet_name="Premissas", index_label="Indicador")
    formatar_planilha(OUTPUT, valores, tipos, unidades)
    print(f"Salvo em {OUTPUT} - indicadores: {list(valores.index)} - colunas: {list(valores.columns)}")


if __name__ == "__main__":
    main()
